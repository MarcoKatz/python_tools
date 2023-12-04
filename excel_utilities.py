# Utilities to format Excel files
# This code initializes a number of styles that are then available as global parameters
# It also has a utility to write a contiguous block of cells 
# In order to usethis, the main code must import the module:   
#       import excel_utilities as ex
# Then in the main code u must invoke the following:
#       if __name__ == "__main__": 
#           ex.set_excel_styles()
#likewise the write_cells function is available as ex.write_cells

from cmath import nan
import numpy as np
import pandas as pd
import os
import shutil
#import xlrd
import re
import time
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def set_excel_styles():

    print("Setting Excel styles")

    global bold_large_style, italic_large_style, bold_underline_large_style, bold_style, cell_style, cell_bold_style, cell_italic_style
    global cell_bold_red_style, cell_small_style, cell_small_style, cell_small_style, cell_align_center_style, cell_align_style
    global cell_align_left_style, cell_align_left_no_wrap_style, cell_border_style, cell_no_border_style, cell_border_merge_style
    global title_border_style, title_align_style, hyperlink_bold_style, hyperlink_small_bold_style, sheet_col_names

    bold_large_style = Font(name='Calibri', size=14, bold=True, italic=False, underline = 'none')
    italic_large_style = Font(name='Calibri', size=14, bold=False, italic=True, underline = 'none')
    bold_underline_large_style = Font(name='Calibri', size=14, bold=True, underline = 'single' )
    bold_style = Font(name='Calibri', size=11, bold=True, italic=False, underline = 'none')
    cell_style = Font(name='Calibri', size=11, bold= False, italic=False, underline = 'none')
    cell_bold_style = Font(name='Calibri', size=11, bold= True, italic=False, underline = 'none')
    cell_italic_style = Font(name='Calibri', size=11, bold= False, italic=True, underline = 'none')
    cell_bold_red_style = Font(name='Calibri', size=11, color = 'FF3300', bold= True, italic=False, underline = 'none')
    cell_small_style = Font(name='Calibri', size=9, bold= False, italic=False, underline = 'none')
#    cell_small_style = Font(name='Calibri', size=9, bold= True, italic=False, underline = 'none')
#    cell_small_style = Font(name='Calibri', size=9, color = 'FF3300', bold= True, italic=False, underline = 'none')
    cell_align_center_style=Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    cell_align_style=Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    cell_align_left_style=Alignment(horizontal='left',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    cell_align_left_no_wrap_style=Alignment(horizontal='left',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)

    cell_border_style = Border(left=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'),\
                top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin',color='FF000000'),\
                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,\
                outline=Side(border_style=None,color='FF000000'), vertical=Side(border_style=None,color='FF000000'),\
                horizontal=Side(border_style=None, color='FF000000'))
    cell_no_border_style = Border(left=Side(border_style=None,color='FF000000'), right=Side(border_style=None,color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None,color='FF000000'),\
                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,\
                outline=Side(border_style=None,color='FF000000'), vertical=Side(border_style=None,color='FF000000'),\
                horizontal=Side(border_style=None, color='FF000000'))
    cell_border_merge_style = Border(left=Side(border_style=None,color='FF000000'), right=Side(border_style=None,color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None,color='FF000000'),\
                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,\
                outline=Side(border_style='thin',color='FF000000'), vertical=Side(border_style=None,color='FF000000'),\
                horizontal=Side(border_style=None, color='FF000000'))
    title_border_style = Border(left=Side(border_style='medium',color='FF000000'), right=Side(border_style='medium',color='FF000000'),\
                top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium',color='FF000000'),\
                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,\
                outline=Side(border_style=None,color='FF000000'), vertical=Side(border_style=None,color='FF000000'),\
                horizontal=Side(border_style=None, color='FF000000'))
    title_align_style=Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=True,indent=0)
    hyperlink_bold_style = Font(name='Calibri', size=11, bold= True, italic=False, underline='single', color='0563C1')
    hyperlink_small_bold_style = Font(name='Calibri', size=9, bold= True, italic=False, underline='single', color='0563C1')
    sheet_col_names = ['A','B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M','N','O','P','Q','R','S','T','U','V',\
                  'W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']

def blank_out_cells(wb_sheet,first_col,first_row, last_col,last_row):
    num_cols = (last_col - first_col + 1)
    num_rows = (last_row - first_row + 1)

    lov_index = -1
    for row in range(first_row,last_row+1):
        for col in range(first_col,last_col+1):
            lov_index += 1
            sheet_pos = sheet_col_names[col-1] + str(row)
            wb_sheet[sheet_pos].value = nan
    
    return wb_sheet


def write_cells(wb_sheet,first_col,first_row, last_col,last_row,font_style,border_style,alignment_style,lov):

    #   purpose:
    #       write a contiguous block of cells 
    #   input: 
    #       wb_sheet: the sheet that is currently open in the Excel document
    #       first_col: the first column from which you write the data: use column 1 to signal Excel column A (and not column 0 in Python)
    #       first_row: the first row from which you write the data: use row 1 to signal Excel row 1 (and not row 0 in Python)
    #       last_col: the first column to which you write the data: starting to count from column 1    
    #       last_row: the first row to which you write the data: starting to count from row 1
    #              the entire matrix size is therefore (last_col - first_col + 1) *  (last_row - first_row+ 1)
    #       font_style: global parameter for appearance of text in cell
    #       border_style:  global parameter for appearance of border
    #       alignment_style:  global parameter for alignment of text in cell
    #       lov: a list of values to be written, which must have a length equal to the matrix size
    #           data must be presented as row by row  (row 1, col 1 col 2 col 3, ...., row 2, col 1 col 2 col 3 )
    #  action: 
    #       
    #  output:
    #       wb_sheet: the sheet that is currently open in the Excel document 
    #  usage example:
    #       wb_sheet : ex.write_cells(wb_sheet, 1, 2, 2, 4, cell_bold_style, cell_border_style, cell_align_left_style, ['Paul',25,'Jim', 12, 'Mary', 41, 'Anne', 22])

    
    num_cols = (last_col - first_col + 1)
    num_rows = (last_row - first_row + 1)

    lov_index = -1
    for row in range(first_row,last_row+1):
        for col in range(first_col,last_col+1):
            lov_index += 1
            sheet_pos = sheet_col_names[col-1] + str(row)
            wb_sheet[sheet_pos].font = font_style
            wb_sheet[sheet_pos].border = border_style
            wb_sheet[sheet_pos].alignment = alignment_style
            wb_sheet[sheet_pos].value = lov[lov_index]
    
    return wb_sheet

def write_excel_file (df_in, file_prod = False,  path_in = "", file_name_in = "", file_ext_in = ["xlsx"] , \
                width_adjust = False, sheet_name_in = 'Sheet1', overwrite = True):
    
    
    if file_prod:
        if file_name_in == "":
#            for x in globals():
#                if globals()[x] is df_in:
#                    print("found name and it is {}".format(x))
#                    print(globals()[x])
            name_df_in =[x for x in globals() if globals()[x] is df_in][0]
            
            file_name_root = path_in + "/" + name_df_in
        else:
            file_name_root = path_in + "/" + file_name_in
        
        for file_ext in file_ext_in:       
            if file_ext not in ['csv','xlsx']:
                continue
            else:
                file_name = file_name_root +'.'+file_ext

            if file_ext == 'csv':
                print("Writing file {}".format(file_name))
                df_in.to_csv(file_name,encoding='UTF-8',index=False)
            else:
                if not overwrite:
                    if os.path.exists(file_name):
                        print("Deleting file {}".format(file_name))
                        os.remove(file_name)
                print("Writing file {}".format(file_name))
                df_in.to_excel(file_name, sheet_name = sheet_name_in, engine = 'openpyxl', index = False, float_format="%.2f")
                wb = openpyxl.load_workbook(file_name)
                ws = wb.active
                if width_adjust :
                    print('Formatting the xlsx file')
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb.active
                    dims = {}
                    for row in ws.rows:
                        for cell in row:
                            if cell.value:
                                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
                        for col, value in dims.items():
                            ws.column_dimensions[col].width = value
                    
#                full_range = "A1:" + get_column_letter(ws.max_column) + str(ws.max_row)
#                ws.auto_filter.ref = full_range
                ws.auto_filter.ref = ws.dimensions
                wb.save(file_name) 


def open_excel_wb(df_in, path_in = "", file_root_name_in = ""):
   
    if file_root_name_in == "":
        name_df_in =[x for x in globals() if globals()[x] is df_in][0]
        file_name_root = path_in + "/" + name_df_in
    else:
        file_name_root = path_in + "/" + file_root_name_in
    
    file_name = file_name_root + ".xlsx"

    if os.path.exists(file_name):
        print('Using file {}'.format(file_name))
    else:
        wb = openpyxl.Workbook()
        wb.save(file_name)
        print('Creating file {}'.format(file_name))
    wb = openpyxl.load_workbook(file_name)
    return wb, file_name


def return_excel_wb(path_in, file_root_name_in):

    file_name = path_in + '/'+ file_root_name_in + '.xlsx'
    wb = openpyxl.load_workbook(file_name)
    return wb


def write_excel_sheet_in_wb(df_in, wb_in, sheet_index_in = 0, sheet_name_in = "no_sheet", width_adjust = False, overwrite = False):

    existing_sheets = wb_in.sheetnames

    if sheet_name_in in existing_sheets:
        if overwrite:
            print('Deleting sheet {}'.format(sheet_name_in))
            sheet_to_del = wb_in.get_sheet_by_name(sheet_name_in)
            wb_in.remove_sheet(sheet_to_del)
        else:
            print('Cannot create sheet {} as it already exists'.format(sheet_name_in))
            return
    
    
    print('Creating sheet {} at index {}'.format(sheet_name_in, sheet_index_in))
    wb_in.create_sheet(index = sheet_index_in, title = sheet_name_in)
    col_names = df_in.columns
    num_cols = len(col_names)
    #print(num_cols)
    
    num_rows = df_in.shape[0]
    
    active_sheet = wb_in[sheet_name_in]

    active_sheet = write_cells(active_sheet,1,1,num_cols,1,cell_bold_style,cell_border_style,cell_align_left_style,\
                        col_names)
    
    for col_index, col in enumerate(col_names):
        active_sheet = write_cells(active_sheet,1+col_index,2,1+col_index,1+num_rows,cell_small_style,cell_no_border_style,\
                            cell_align_left_style,df_in[col].values)

    if width_adjust :
        dims = {}
        for row in active_sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
            for col, value in dims.items():
                active_sheet.column_dimensions[col].width = min(value,50)

    active_sheet.auto_filter.ref = active_sheet.dimensions
    active_sheet.freeze_panes = 'A2'


def write_excel_sheet (df_in, file_prod = False,  path_in = "", file_root_name_in = "", sheet_index_in = 0, sheet_name_in = "no_sheet", width_adjust = False, overwrite = False):

    if file_prod:        
        wb, file_name = open_excel_wb(df_in, path_in, file_root_name_in)

        write_excel_sheet_in_wb(df_in, wb, sheet_index_in, sheet_name_in, width_adjust, overwrite)

        wb.save(file_name)